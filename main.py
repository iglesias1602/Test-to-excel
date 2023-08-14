import openpyxl
import unittest
import sys
import importlib
import io
import contextlib


# Custom TestResult class to capture test results
class CustomTestResult(unittest.TextTestResult):
    def __init__(self, stream=sys.stdout, descriptions=1, verbosity=1):
        super().__init__(stream, descriptions, verbosity)
        self.test_results = []

    def addSuccess(self, test):
        super().addSuccess(test)
        test_class = test.__class__.__name__
        test_method = test._testMethodName
        input_data = getattr(test, "input_data", "")
        self.test_results.append((test_class, test_method, input_data, "OK"))

    def addFailure(self, test, err):
        super().addFailure(test, err)
        test_class = test.__class__.__name__
        test_method = test._testMethodName
        input_data = getattr(test, "input_data", "")
        self.test_results.append((test_class, test_method, input_data, "NOT OK"))


# Run tests and capture results
def run_tests(test_module):
    test_suite = unittest.defaultTestLoader.loadTestsFromModule(test_module)
    result = CustomTestResult()
    test_suite.run(result)
    return result.test_results


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print(
            "Usage: python run_tests_and_generate_excel.py <test_file.py> <output_file.xlsx>"
        )
        sys.exit(1)

    test_file = sys.argv[1]
    output_file = sys.argv[2]

    try:
        module = importlib.import_module(test_file.replace(".py", ""))
    except ImportError:
        print(f"Error: Failed to import test file '{test_file}'")
        sys.exit(1)

    test_classes = [
        cls
        for cls in module.__dict__.values()
        if isinstance(cls, type) and issubclass(cls, unittest.TestCase)
    ]
    test_results = []

    # Loop through test classes and their methods
    for test_class in test_classes:
        test_suite = unittest.TestLoader().loadTestsFromTestCase(test_class)

        for test in test_suite:
            # Capture the input_data attribute from each test method
            input_data = getattr(test, "input_data", "")

            # Redirect stdout to capture the print output from the test
            stdout_backup = sys.stdout
            captured_output = io.StringIO()
            sys.stdout = captured_output

            # Run the test and capture the printed output
            with contextlib.redirect_stdout(captured_output):
                test_result = unittest.TextTestRunner(
                    resultclass=CustomTestResult, stream=captured_output
                ).run(test)

            # Restore stdout
            sys.stdout = stdout_backup

            # Determine test result (OK/NOT OK)
            if test_result.wasSuccessful():
                result = "OK"
            else:
                result = "NOT OK"

            # Append test results
            test_results.append(
                (
                    test_class.__name__,
                    test._testMethodName,
                    input_data,
                    result,
                    captured_output.getvalue(),
                )
            )

    # Write test results to Excel
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Test Results"

    sheet["A1"] = "Test Class"
    sheet["B1"] = "Test Method"
    sheet["C1"] = "Input"
    sheet["D1"] = "Result"
    sheet["E1"] = "Print Output"

    for row, (test_class, test_method, input_data, result, print_output) in enumerate(
        test_results, start=2
    ):
        sheet.cell(row=row, column=1, value=test_class)
        sheet.cell(row=row, column=2, value=test_method)
        sheet.cell(row=row, column=3, value=input_data)
        sheet.cell(row=row, column=4, value=result)
        sheet.cell(
            row=row, column=5, value=print_output.strip()
        )  # Strip newline from captured output

    wb.save(output_file)

    print(f"Test results generated and written to {output_file}")
